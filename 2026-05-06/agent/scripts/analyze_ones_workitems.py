#!/usr/bin/env python3
import csv
import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


PRIORITY_MAP = {"P0": "最高", "P1": "较高", "P2": "普通", "P3": "较低"}


def normalize(value):
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    return re.sub(r"[\s\W_]+", "", text, flags=re.UNICODE)


def first_line(value):
    lines = str(value or "").strip().splitlines()
    return lines[0].strip() if lines else ""


def header_map(ws):
    headers = {}
    for idx, value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), start=1):
        if value:
            headers[str(value).strip()] = idx
    return headers


def cell(row, headers, name):
    idx = headers.get(name)
    if not idx:
        return None
    return row[idx - 1]


def read_existing_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return set()
    ws = wb[sheet_name]
    headers = header_map(ws)
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        content = cell(row, headers, "具体需求内容")
        if content:
            existing.add(normalize(content))
    return existing


def parse_existing_text(text):
    values = set()
    for line in str(text or "").splitlines():
        line = line.strip()
        if not line:
            continue
        line = re.sub(r"^#?\d+\s*", "", line)
        if "_" in line:
            line = line.split("_", 1)[1]
        values.add(normalize(line))
    return values


def analyze(payload):
    workbook = payload["spreadsheetPath"]
    target_iteration = payload["targetIteration"]
    assignee = payload.get("assignee") or "wangpeng5@tetras.ai"
    existing_target = parse_existing_text(payload.get("targetExistingItemsText"))

    wb = load_workbook(workbook, data_only=True)
    ws = wb["需求汇总"]
    headers = header_map(ws)
    v02_existing = read_existing_sheet(wb, "v0.2")

    required = ["需求分类", "具体需求内容", "优先级"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"需求汇总缺少字段：{', '.join(missing)}")

    category = ""
    valid = []
    excluded = []
    seen = set()
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_category = cell(row, headers, "需求分类")
        if raw_category:
            category = first_line(raw_category)
        content = cell(row, headers, "具体需求内容")
        priority = first_line(cell(row, headers, "优先级"))
        status = first_line(cell(row, headers, "状态"))
        if not content or not priority or not category:
            continue

        key = normalize(content)
        item = {
            "row": row_number,
            "category": category,
            "content": str(content).strip(),
            "priority": priority,
            "status": status or "",
            "title": f"【{priority}】{category}_{str(content).strip()}",
        }

        if status == "已实现":
            excluded.append({**item, "reason": "状态=已实现"})
        elif key in v02_existing:
            excluded.append({**item, "reason": "已存在于 v0.2"})
        elif key in existing_target:
            excluded.append({**item, "reason": "目标迭代已存在相同需求"})
        elif key in seen:
            excluded.append({**item, "reason": "需求汇总内重复"})
        else:
            seen.add(key)
            valid.append(item)

    import_rows = []
    for item in valid:
        import_rows.append({
            "标题": item["title"],
            "工作项类型": "需求",
            "负责人": assignee,
            "状态": "未开始",
            "所属项目": "Tetrasphere产品开发",
            "所属迭代": target_iteration,
            "优先级": PRIORITY_MAP.get(item["priority"], item["priority"]),
        })

    return {
        "totalValidSourceRows": len(valid) + len(excluded),
        "pendingCreateCount": len(valid),
        "excludedCount": len(excluded),
        "pending": valid,
        "excluded": excluded,
        "importRows": import_rows,
    }


def main():
    if len(sys.argv) != 4:
        raise SystemExit("Usage: analyze_ones_workitems.py <payload.json> <output.json> <output.csv>")
    payload = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    result = analyze(payload)
    Path(sys.argv[2]).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    with Path(sys.argv[3]).open("w", newline="", encoding="utf-8-sig") as f:
        fieldnames = ["标题", "工作项类型", "负责人", "状态", "所属项目", "所属迭代", "优先级"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(result["importRows"])
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
